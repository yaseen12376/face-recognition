import json
import os
from datetime import datetime
from tabulate import tabulate
import pandas as pd

LOCAL_ATTENDANCE_LOG = "local_attendance.json"

def create_beautiful_attendance_table():
    """Create a beautiful table view of attendance data with First Arrival and Latest Visit"""
    
    if not os.path.exists(LOCAL_ATTENDANCE_LOG):
        print("❌ No attendance records found yet.")
        print("🎯 Run the face recognition system first!")
        return
    
    with open(LOCAL_ATTENDANCE_LOG, 'r') as f:
        data = json.load(f)
    
    if not data:
        print("📊 No attendance data available.")
        return
    
    # Get all students
    all_students = set()
    for date_data in data.values():
        all_students.update(date_data.keys())
    
    all_students = sorted(list(all_students))
    
    # Check if we have new format (with first_arrival/latest_visit) or old format
    sample_data = next(iter(data.values())) if data else {}
    sample_student = next(iter(sample_data.values())) if sample_data else {}
    new_format = isinstance(sample_student, dict) and 'first_arrival' in sample_student
    
    if new_format:
        # Create table with First Arrival and Latest Visit columns
        table_data = []
        headers = ["Date"]
        for student in all_students:
            headers.extend([f"{student}_First", f"{student}_Latest"])
        
        for date in sorted(data.keys(), reverse=True):
            row = [date]
            for student in all_students:
                student_data = data[date].get(student, {})
                if isinstance(student_data, dict):
                    first = student_data.get('first_arrival', '')
                    latest = student_data.get('latest_visit', '')
                    row.extend([first, latest])
                else:
                    # Old format compatibility
                    row.extend([student_data, student_data])
            table_data.append(row)
            
        # Display beautiful table
        print("\n" + "="*100)
        print("📊 FACE RECOGNITION ATTENDANCE RECORDS")
        print("🎯 First Arrival | 🔄 Latest Visit")
        print("="*100)
        
    else:
        # Old format - single time per student
        table_data = []
        headers = ["Date"] + all_students
        
        for date in sorted(data.keys(), reverse=True):
            row = [date]
            for student in all_students:
                time_val = data[date].get(student, "")
                row.append(time_val)
            table_data.append(row)
            
        print("\n" + "="*80)
        print("📊 FACE RECOGNITION ATTENDANCE RECORDS")
        print("="*80)
    
    # Use tabulate for beautiful formatting
    try:
        print(tabulate(table_data, headers=headers, tablefmt="grid", stralign="center"))
    except:
        # Fallback if tabulate not available
        print(tabulate(table_data, headers=headers, tablefmt="simple"))
    
    print("="*100 if new_format else "="*80)
    
    # Summary statistics
    total_days = len(data)
    today = datetime.now().strftime("%Y-%m-%d")
    today_count = len(data.get(today, {}))
    
    print(f"📈 Summary:")
    print(f"   📅 Total days recorded: {total_days}")
    print(f"   👥 Students registered: {len(all_students)}")
    print(f"   🎯 Today's attendance: {today_count} students")
    
    if new_format:
        print(f"   🎯 First Arrival: Records when student first enters")
        print(f"   🔄 Latest Visit: Updates every time student is detected")
    
    # Most active students
    student_counts = {}
    for date_data in data.values():
        for student in date_data:
            student_counts[student] = student_counts.get(student, 0) + 1
    
    if student_counts:
        most_active = max(student_counts, key=student_counts.get)
        print(f"   🏆 Most active student: {most_active} ({student_counts[most_active]} days)")
    
    print("="*100 if new_format else "="*80)

def export_to_excel():
    """Export attendance data to Excel with beautiful multi-level headers"""
    if not os.path.exists(LOCAL_ATTENDANCE_LOG):
        print("❌ No attendance data to export.")
        return
    
    with open(LOCAL_ATTENDANCE_LOG, 'r') as f:
        data = json.load(f)
    
    # Get all students
    all_students = set()
    for date_data in data.values():
        all_students.update(date_data.keys())
    
    all_students = sorted(list(all_students))
    
    # Check format
    sample_data = next(iter(data.values())) if data else {}
    sample_student = next(iter(sample_data.values())) if sample_data else {}
    new_format = isinstance(sample_student, dict) and 'first_arrival' in sample_student
    
    # Create Excel file with beautiful formatting
    excel_file = "attendance_backup.xlsx"
    
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Face Recognition Attendance"
        
        if new_format:
            # Create beautiful multi-level headers
            # Row 1: Student Names (merged across 2 columns each)
            ws.cell(row=1, column=1, value="DATE")
            ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
            
            col = 2
            for student in all_students:
                ws.cell(row=1, column=col, value=student.upper())
                ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col+1)
                col += 2
            
            # Row 2: First Arrival | Latest Visit
            col = 2
            for student in all_students:
                ws.cell(row=2, column=col, value="First Arrival")
                ws.cell(row=2, column=col+1, value="Latest Visit")
                col += 2
            
            # Add data rows
            row_num = 3
            for date in sorted(data.keys()):
                ws.cell(row=row_num, column=1, value=date)
                col = 2
                for student in all_students:
                    student_data = data[date].get(student, {})
                    if isinstance(student_data, dict):
                        ws.cell(row=row_num, column=col, value=student_data.get('first_arrival', ''))
                        ws.cell(row=row_num, column=col+1, value=student_data.get('latest_visit', ''))
                    else:
                        ws.cell(row=row_num, column=col, value=student_data)
                        ws.cell(row=row_num, column=col+1, value=student_data)
                    col += 2
                row_num += 1
            
            # Apply beautiful formatting
            # Header styling
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            subheader_fill = PatternFill(start_color="8DB4E2", end_color="8DB4E2", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True, size=12)
            subheader_font = Font(color="000000", bold=True, size=10)
            
            # Style main headers (student names)
            for col in range(1, len(all_students) * 2 + 2):
                cell = ws.cell(row=1, column=col)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Style sub-headers (First Arrival | Latest Visit)
            for col in range(1, len(all_students) * 2 + 2):
                cell = ws.cell(row=2, column=col)
                cell.fill = subheader_fill
                cell.font = subheader_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Auto-adjust column widths
            for col in range(1, len(all_students) * 2 + 2):
                ws.column_dimensions[get_column_letter(col)].width = 12
            
            ws.column_dimensions['A'].width = 15  # Date column wider
            
        else:
            # Fallback to simple format for old data
            df_data = []
            for date in sorted(data.keys()):
                row = {"Date": date}
                for student in all_students:
                    row[student] = data[date].get(student, "")
                df_data.append(row)
            
            df = pd.DataFrame(df_data)
            df.to_excel(excel_file, index=False)
            wb = None
        
        if wb:
            wb.save(excel_file)
        
        if new_format:
            print(f"✅ Beautiful attendance exported to: {excel_file}")
            print(f"📊 Excel format:")
            print(f"   🎯 Multi-level headers with student names")
            print(f"   � {len(all_students)} students × 2 columns each (First Arrival | Latest Visit)")
            print(f"   🎨 Professional formatting with colors and borders")
            print(f"📁 Open {excel_file} in Excel to see the beautiful layout!")
        else:
            print(f"✅ Attendance exported to: {excel_file}")
            
    except ImportError:
        # Fallback if openpyxl not available
        df_data = []
        if new_format:
            for date in sorted(data.keys()):
                row = {"Date": date}
                for student in all_students:
                    student_data = data[date].get(student, {})
                    if isinstance(student_data, dict):
                        row[f"{student}_First"] = student_data.get('first_arrival', '')
                        row[f"{student}_Latest"] = student_data.get('latest_visit', '')
                    else:
                        row[f"{student}_First"] = student_data
                        row[f"{student}_Latest"] = student_data
                df_data.append(row)
        else:
            for date in sorted(data.keys()):
                row = {"Date": date}
                for student in all_students:
                    row[student] = data[date].get(student, "")
                df_data.append(row)
        
        df = pd.DataFrame(df_data)
        df.to_excel(excel_file, index=False)
        print(f"✅ Attendance exported to: {excel_file}")
        print("ℹ️  Install openpyxl for beautiful formatting: pip install openpyxl")

def main():
    print("🎯 Beautiful Attendance Viewer")
    print("=" * 50)
    
    while True:
        print("\nOptions:")
        print("1. 📊 View beautiful table")
        print("2. 📁 Export to Excel")
        print("3. 🚪 Exit")
        
        choice = input("\nEnter choice (1-3): ").strip()
        
        if choice == "1":
            create_beautiful_attendance_table()
        elif choice == "2":
            export_to_excel()
        elif choice == "3":
            print("👋 Goodbye!")
            break
        else:
            print("❌ Invalid choice. Please try again.")

if __name__ == "__main__":
    main()