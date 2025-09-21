"""
Web Dashboard for Face Recognition Attendance
Accessible via browser by anyone on the network
"""
from flask import Flask, render_template, jsonify, send_file
import json
import os
from datetime import datetime

app = Flask(__name__)
LOCAL_ATTENDANCE_LOG = "local_attendance.json"

@app.route('/')
def dashboard():
    """Main dashboard page"""
    return render_template('dashboard.html')

@app.route('/api/attendance')
def get_attendance_data():
    """API endpoint to get attendance data"""
    if not os.path.exists(LOCAL_ATTENDANCE_LOG):
        return jsonify({"error": "No attendance data found"})
    
    with open(LOCAL_ATTENDANCE_LOG, 'r') as f:
        data = json.load(f)
    
    # Convert to format suitable for web display
    formatted_data = []
    all_students = set()
    
    for date_data in data.values():
        all_students.update(date_data.keys())
    
    all_students = sorted(list(all_students))
    
    for date in sorted(data.keys(), reverse=True):
        row = {"date": date, "students": {}}
        for student in all_students:
            student_data = data[date].get(student, {})
            if isinstance(student_data, dict):
                row["students"][student] = {
                    "first_arrival": student_data.get('first_arrival', ''),
                    "latest_visit": student_data.get('latest_visit', '')
                }
            else:
                row["students"][student] = {
                    "first_arrival": student_data,
                    "latest_visit": student_data
                }
        formatted_data.append(row)
    
    return jsonify({
        "data": formatted_data,
        "students": all_students,
        "total_days": len(data),
        "last_updated": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    })

@app.route('/api/export')
def export_excel():
    """Generate and download Excel file with beautiful formatting"""
    try:
        if not os.path.exists(LOCAL_ATTENDANCE_LOG):
            return jsonify({"error": "No attendance data found"}), 404
        
        with open(LOCAL_ATTENDANCE_LOG, 'r') as f:
            data = json.load(f)
        
        # Get all students
        all_students = set()
        for date_data in data.values():
            all_students.update(date_data.keys())
        all_students = sorted(list(all_students))
        
        # Create Excel file with beautiful formatting
        excel_file = "web_attendance_export.xlsx"
        
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Face Recognition Attendance"
            
            # Create beautiful multi-level headers
            # Row 1: Student Names (merged across 2 columns each)
            ws.cell(row=1, column=1, value="DATE")
            ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
            
            col = 2
            for student in all_students:
                ws.cell(row=1, column=col, value=student.upper())
                ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col+1)
                col += 2
            
            # Row 2: First Arrival | Latest Visit
            col = 2
            for student in all_students:
                ws.cell(row=2, column=col, value="First Arrival")
                ws.cell(row=2, column=col+1, value="Latest Visit")
                col += 2
            
            # Add data rows
            row_num = 3
            for date in sorted(data.keys(), reverse=True):  # Most recent first
                ws.cell(row=row_num, column=1, value=date)
                col = 2
                for student in all_students:
                    student_data = data[date].get(student, {})
                    if isinstance(student_data, dict):
                        ws.cell(row=row_num, column=col, value=student_data.get('first_arrival', ''))
                        ws.cell(row=row_num, column=col+1, value=student_data.get('latest_visit', ''))
                    else:
                        ws.cell(row=row_num, column=col, value=student_data)
                        ws.cell(row=row_num, column=col+1, value=student_data)
                    col += 2
                row_num += 1
            
            # Apply beautiful formatting
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            subheader_fill = PatternFill(start_color="8DB4E2", end_color="8DB4E2", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True, size=12)
            subheader_font = Font(color="000000", bold=True, size=10)
            
            # Style headers
            for col in range(1, len(all_students) * 2 + 2):
                ws.cell(row=1, column=col).fill = header_fill
                ws.cell(row=1, column=col).font = header_font
                ws.cell(row=1, column=col).alignment = Alignment(horizontal="center", vertical="center")
                
                ws.cell(row=2, column=col).fill = subheader_fill
                ws.cell(row=2, column=col).font = subheader_font
                ws.cell(row=2, column=col).alignment = Alignment(horizontal="center", vertical="center")
            
            # Auto-adjust column widths
            for col in range(1, len(all_students) * 2 + 2):
                ws.column_dimensions[get_column_letter(col)].width = 12
            ws.column_dimensions['A'].width = 15
            
            wb.save(excel_file)
            
            return send_file(excel_file, as_attachment=True, 
                           download_name=f"attendance_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
                           
        except ImportError:
            # Fallback to pandas if openpyxl not available
            import pandas as pd
            
            df_data = []
            for date in sorted(data.keys(), reverse=True):
                row = {"Date": date}
                for student in all_students:
                    student_data = data[date].get(student, {})
                    if isinstance(student_data, dict):
                        row[f"{student}_First"] = student_data.get('first_arrival', '')
                        row[f"{student}_Latest"] = student_data.get('latest_visit', '')
                    else:
                        row[f"{student}_First"] = student_data
                        row[f"{student}_Latest"] = student_data
                df_data.append(row)
            
            df = pd.DataFrame(df_data)
            excel_file = f"attendance_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            df.to_excel(excel_file, index=False)
            
            return send_file(excel_file, as_attachment=True)
            
    except Exception as e:
        return jsonify({"error": f"Failed to generate Excel: {str(e)}"}), 500

if __name__ == '__main__':
    # Create templates folder if it doesn't exist
    if not os.path.exists('templates'):
        os.makedirs('templates')
    
    print("🌐 Starting Face Recognition Attendance Web Dashboard")
    print("=" * 60)
    print("📊 Dashboard will be available at:")
    print("   🖥️  Local: http://localhost:5000")
    print("   📱 Network: http://[YOUR-IP]:5000")
    print("=" * 60)
    print("🔗 Share the network URL with your colleagues!")
    print("✨ They can view attendance in real-time via browser")
    
    app.run(host='0.0.0.0', port=5000, debug=False)